#!/usr/bin/env python3
"""
PF2e DB 관리 도구
- export: JS DB → Excel (.xlsx)
- import: Excel (.xlsx) → JS DB

사용법:
  python3 db_tools.py export          # DB → Excel 내보내기
  python3 db_tools.py import          # Excel → DB 가져오기
  python3 db_tools.py import --dry    # 미리보기 (파일 변경 없음)
"""

import sys, os, json, re, subprocess
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl 필요: pip3 install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
EXCEL_FILE = SCRIPT_DIR / "PF2e_DB.xlsx"
JSON_CACHE = SCRIPT_DIR / "_db_cache.json"

# ═══════════════════════════════════════
#  JS ↔ JSON 변환
# ═══════════════════════════════════════

def load_dbs_via_node():
    """Node.js로 JS DB 파일을 eval해서 JSON 캐시 생성"""
    result = subprocess.run(['node', str(SCRIPT_DIR / 'db_export.js')],
                          capture_output=True, text=True, cwd=str(SCRIPT_DIR))
    if result.returncode != 0:
        print(f"❌ Node.js 실행 실패:\n{result.stderr}")
        sys.exit(1)
    print(result.stdout.strip())
    return json.loads(JSON_CACHE.read_text(encoding='utf-8'))


def array_to_js(var_name, data, fields, declaration='const'):
    """배열 데이터를 JS 코드로 변환"""
    lines = []
    for item in data:
        parts = []
        for f in fields:
            val = item.get(f)
            if val is None:
                parts.append(f"{f}:null")
            elif isinstance(val, bool):
                parts.append(f"{f}:{'true' if val else 'false'}")
            elif isinstance(val, (int, float)):
                if val == int(val):
                    parts.append(f"{f}:{int(val)}")
                else:
                    parts.append(f"{f}:{val}")
            elif isinstance(val, list):
                inner = ','.join(f"'{v}'" for v in val)
                parts.append(f"{f}:[{inner}]")
            else:
                escaped = str(val).replace("\\", "\\\\").replace("'", "\\'")
                parts.append(f"{f}:'{escaped}'")
        lines.append('  {' + ', '.join(parts) + '},')
    return f"{declaration} {var_name} = [\n" + '\n'.join(lines) + "\n];"


def dict_to_js(var_name, data, declaration='var'):
    """dict(key→array) 데이터를 JS 코드로 변환 (CLASS_AUTO_FEATS 등)"""
    lines = [f"{declaration} {var_name} = {{"]
    for key, arr in data.items():
        if not arr:
            continue
        items = []
        for item in arr:
            parts = []
            for k, v in item.items():
                if v is None:
                    parts.append(f"{k}:null")
                elif isinstance(v, bool):
                    parts.append(f"{k}:{'true' if v else 'false'}")
                elif isinstance(v, (int, float)):
                    parts.append(f"{k}:{int(v) if v == int(v) else v}")
                else:
                    escaped = str(v).replace("\\", "\\\\").replace("'", "\\'")
                    parts.append(f"{k}:'{escaped}'")
            items.append('{' + ', '.join(parts) + '}')
        lines.append(f"  '{key}': [{', '.join(items)}],")
    lines.append("};")
    return '\n'.join(lines)


# ═══════════════════════════════════════
#  EXPORT: JS → Excel
# ═══════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1a2332", end_color="1a2332", fill_type="solid")
HEADER_FONT = Font(color="e67e22", bold=True, size=11)
CELL_FONT = Font(color="333333", size=10)

def write_sheet(wb, sheet_name, data, fields):
    ws = wb.create_sheet(sheet_name)
    for col, f in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col, value=f)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    for row_idx, item in enumerate(data, 2):
        for col, f in enumerate(fields, 1):
            val = item.get(f)
            if isinstance(val, list):
                val = ', '.join(str(v) for v in val)
            elif isinstance(val, dict):
                val = json.dumps(val, ensure_ascii=False)
            elif isinstance(val, bool):
                val = 'TRUE' if val else 'FALSE'
            ws.cell(row=row_idx, column=col, value=val).font = CELL_FONT
    # Auto-width
    for col in range(1, len(fields)+1):
        letter = openpyxl.utils.get_column_letter(col)
        max_len = 0
        for r in range(1, min(len(data)+2, 100)):
            cell_val = str(ws.cell(row=r, column=col).value or '')
            max_len = max(max_len, min(len(cell_val), 50))
        ws.column_dimensions[letter].width = max_len + 4
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    print(f"  ✓ {sheet_name}: {len(data)}개 항목")


def write_dict_sheet(wb, sheet_name, data, fields):
    """dict(key→[items]) 데이터를 시트로 (key 열 추가)"""
    flat = []
    for key, arr in data.items():
        for item in arr:
            row = {'_key': key}
            row.update(item)
            flat.append(row)
    all_fields = ['_key'] + fields
    write_sheet(wb, sheet_name, flat, all_fields)


def do_export():
    print("═══ DB → Excel 내보내기 ═══\n")
    dbs = load_dbs_via_node()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 기존 시트 ──
    write_sheet(wb, '주문_SPELL', dbs['SPELL_DB'],
        ['name_ko','name_en','rank','is_cantrip','is_focus','traditions','actions','traits','summary','desc'])
    write_sheet(wb, '재주_FEAT', dbs['FEAT_DB'],
        ['name_ko','name_en','feat_level','prerequisites','traits','category','summary','desc'])
    write_sheet(wb, '무기_WEAPON', dbs['WEAPON_DB'],
        ['name_ko','name_en','category','price','damage','bulk','hands','range','reload','group','traits'])
    write_sheet(wb, '갑옷_ARMOR', dbs['ARMOR_DB'],
        ['name_ko','name_en','category','price','ac_bonus','dex_cap','check_penalty','speed_penalty','strength','bulk','group','traits'])
    write_sheet(wb, '방패_SHIELD', dbs['SHIELD_DB'],
        ['name_ko','name_en','price','ac_bonus','speed_penalty','bulk','hardness','hp','bt'])
    write_sheet(wb, '장비_GEAR', dbs['GEAR_DB'],
        ['name_ko','name_en','price','bulk'])

    # ── 캐릭터 기본 데이터 ──
    write_sheet(wb, '클래스_CLASS', dbs['CLASSES'],
        ['id','name','en','hp','keyAttr','tradition','casting','skills','desc'])
    write_sheet(wb, '서브클래스_SUBCLASS', dbs['SUBCLASS_DB'],
        ['id','class_id','subclass_type','name_ko','name_en','summary'])
    write_sheet(wb, '혈통_ANCESTRY', dbs['ANCESTRIES'],
        ['id','name','en','hp','size','speed','boosts','flaws','traits','desc'])
    write_sheet(wb, '배경_BACKGROUND', dbs['BACKGROUNDS'],
        ['id','name','en','boosts','skills','feat','desc'])
    write_sheet(wb, '유산_HERITAGE', dbs['HERITAGE_DB'],
        ['id','name_ko','name_en','ancestry','summary'])

    # ── 신격/동료 ──
    write_sheet(wb, '신격_DEITY', dbs['DEITY_DB'],
        ['id','name_ko','name_en','weapon','skill','sanctification','domains'])
    write_sheet(wb, '동물동료_COMPANION', dbs['COMPANION_DB'],
        ['id','name_ko','name_en','size','hp','str','dex','con','int','wis','cha',
         'skill','senses','speed','speeds','mount','uncommon','attacks'])
    write_sheet(wb, '마갑_BARDING', dbs['BARDING_DB'],
        ['name','ac','dex','check','speed','bulk','category'])
    write_sheet(wb, '사역마능력_FAMILIAR', dbs['FAMILIAR_ABILITIES'],
        ['id','name','en','desc'])

    # ── 상태이상 ──
    write_sheet(wb, '상태이상_CONDITION', dbs['CONDITIONS_DATA'],
        ['name','en','valued','max','desc'])

    # ── 행동 ──
    write_sheet(wb, '행동_ACTION', dbs['ACTION_DB'],
        ['id','cat','cat_label','name_ko','name_en','cost','traits','req_skill','req_rank','req_feat','summary'])

    # ── 기술 ──
    write_sheet(wb, '기술_SKILL', dbs['SKILLS'],
        ['id','name','en','attr','isLore'])

    # ── 클래스 요소 (dict 형태) ──
    write_dict_sheet(wb, '클래스요소_FEATURE', dbs['CLASS_FEATURE_NAMES'],
        ['lv','name_ko','name_en','type'])
    write_dict_sheet(wb, '서브클래스요소_SUBFEATURE', dbs['SUBCLASS_FEATURE_NAMES'],
        ['lv','name_ko','name_en'])
    write_dict_sheet(wb, '자동재주_AUTOFEAT', dbs['CLASS_AUTO_FEATS'],
        ['lv','name_ko','name_en','category'])
    write_dict_sheet(wb, '서브자동재주_SUBAUTOFEAT', dbs['SUBCLASS_AUTO_FEATS'],
        ['lv','name_ko','name_en','category'])
    write_dict_sheet(wb, '자동주문_AUTOSPELL', dbs['CLASS_AUTO_SPELLS'],
        ['lv','type','name_ko','name_en'])
    write_dict_sheet(wb, '서브자동주문_SUBAUTOSPELL', dbs['SUBCLASS_AUTO_SPELLS'],
        ['lv','type','name_ko','name_en'])

    wb.save(EXCEL_FILE)
    print(f"\n✅ 저장: {EXCEL_FILE}")
    print(f"   수정 후 → python3 db_tools.py import")


# ═══════════════════════════════════════
#  IMPORT: Excel → JS
# ═══════════════════════════════════════

# 숫자로 변환할 필드
INT_FIELDS = {'rank','feat_level','ac_bonus','check_penalty','speed_penalty',
              'strength','hands','range','reload','hardness','hp','bt',
              'lv','ac','dex','check','speed','bulk','max','req_rank',
              'str','con','int','wis','cha'}
BOOL_FIELDS = {'is_cantrip','is_focus','valued','mount','uncommon','isLore'}
LIST_FIELDS = {'traits','traditions','sanctification','domains','boosts','flaws'}

def read_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ 시트 '{sheet_name}' 없음")
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2: return []
    headers = [str(h).strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row): continue
        item = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if h in BOOL_FIELDS:
                if val == 'TRUE' or val is True:
                    val = True
                elif val == 'FALSE' or val is False:
                    val = False
                else:
                    val = bool(val) if val else False
            elif h in LIST_FIELDS:
                val = [v.strip() for v in str(val).split(',') if v.strip()] if val and str(val).strip() not in ('', 'None') else []
            elif h in INT_FIELDS:
                if val is not None and str(val).strip() not in ('', '—', '-', 'None'):
                    try: val = int(float(val))
                    except: pass
                else: val = None
            elif h == 'dex_cap':
                if val is not None and str(val).strip() not in ('', '—', 'None'):
                    try: val = int(float(val))
                    except: val = None
                else: val = None
            elif h == 'bulk':
                sv = str(val).strip() if val else ''
                if sv in ('L','l'): val = 'L'
                elif sv in ('—','-','','None'): val = '—'
                else:
                    try: val = int(float(sv))
                    except: pass
            elif h in ('speeds','attacks'):
                # JSON 직렬화된 dict/list
                if val and str(val).strip() not in ('', 'None', '{}', '[]'):
                    try: val = json.loads(str(val))
                    except: pass
                else:
                    val = {} if h == 'speeds' else []
            elif val is None:
                val = ''
            item[h] = val
        data.append(item)
    print(f"  ✓ {sheet_name}: {len(data)}개")
    return data


def read_dict_sheet(wb, sheet_name):
    """_key 열이 있는 시트를 dict(key→[items])로 변환"""
    rows = read_sheet(wb, sheet_name)
    if rows is None: return None
    result = {}
    for row in rows:
        key = row.pop('_key', '')
        if not key: continue
        # None/빈문자열 필드 정리
        cleaned = {}
        for k, v in row.items():
            if v is not None and v != '':
                cleaned[k] = v
        result.setdefault(key, []).append(cleaned)
    return result


def do_import(dry_run=False):
    print("═══ Excel → DB 가져오기 ═══\n")
    if not EXCEL_FILE.exists():
        print(f"❌ {EXCEL_FILE} 없음. export 먼저 실행하세요.")
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # ═══ SPELL_DB ═══
    spells = read_sheet(wb, '주문_SPELL')
    if spells is not None:
        js = "// Pathfinder 2e Player Core — 주문 데이터베이스\n// Excel → db_tools.py import\n\n"
        js += array_to_js('SPELL_DB', spells,
            ['name_ko','name_en','rank','is_cantrip','is_focus','traditions','actions','traits','summary','desc'])
        if not dry_run: (SCRIPT_DIR/'SPELL_DB.js').write_text(js, encoding='utf-8')
        print(f"    → SPELL_DB.js {'(미리보기)' if dry_run else '✓'}")

    # ═══ FEAT_DB ═══
    feats = read_sheet(wb, '재주_FEAT')
    if feats is not None:
        js = "// Pathfinder 2e Player Core — Feat Database\n// Excel → db_tools.py import\n\n"
        js += array_to_js('FEAT_DB', feats,
            ['name_ko','name_en','feat_level','prerequisites','traits','category','summary','desc'], 'const')
        if not dry_run: (SCRIPT_DIR/'feat_db.js').write_text(js, encoding='utf-8')
        print(f"    → feat_db.js {'(미리보기)' if dry_run else '✓'}")

    # ═══ Equipment (무기/갑옷/방패/장비) ═══
    weapons = read_sheet(wb, '무기_WEAPON')
    armors = read_sheet(wb, '갑옷_ARMOR')
    shields = read_sheet(wb, '방패_SHIELD')
    gears = read_sheet(wb, '장비_GEAR')
    js = "// Pathfinder 2e Player Core — Equipment Database\n// Excel → db_tools.py import\n\n"
    if armors is not None:
        js += array_to_js('ARMOR_DB', armors,
            ['name_ko','name_en','category','price','ac_bonus','dex_cap','check_penalty','speed_penalty','strength','bulk','group','traits']) + "\n\n"
    if shields is not None:
        js += array_to_js('SHIELD_DB', shields,
            ['name_ko','name_en','price','ac_bonus','speed_penalty','bulk','hardness','hp','bt']) + "\n\n"
    if weapons is not None:
        js += array_to_js('WEAPON_DB', weapons,
            ['name_ko','name_en','category','price','damage','bulk','hands','range','reload','group','traits']) + "\n\n"
    if gears is not None:
        js += array_to_js('GEAR_DB', gears, ['name_ko','name_en','price','bulk']) + "\n"
    if not dry_run: (SCRIPT_DIR/'equipment_db.js').write_text(js, encoding='utf-8')
    print(f"    → equipment_db.js {'(미리보기)' if dry_run else '✓'}")

    # ═══ cs_data.js ═══
    classes = read_sheet(wb, '클래스_CLASS')
    subclasses = read_sheet(wb, '서브클래스_SUBCLASS')
    ancestries = read_sheet(wb, '혈통_ANCESTRY')
    backgrounds = read_sheet(wb, '배경_BACKGROUND')
    heritages = read_sheet(wb, '유산_HERITAGE')
    conditions_data = read_sheet(wb, '상태이상_CONDITION')
    skills = read_sheet(wb, '기술_SKILL')
    actions = read_sheet(wb, '행동_ACTION')

    if any(x is not None for x in [classes, subclasses, ancestries, backgrounds, heritages, conditions_data, skills, actions]):
        # 기존 cs_data.js 읽기 — 덮어쓸 부분만 교체
        cs_data_path = SCRIPT_DIR / 'cs_data.js'
        orig = cs_data_path.read_text(encoding='utf-8')

        js_parts = ["// ═══════════════════════════════════════════════\n//  DATA — from Player Core Korean translation\n// ═══════════════════════════════════════════════\n"]

        if classes is not None:
            # saves 필드가 문자열이면 패스스루
            js_parts.append(array_to_js('CLASSES', classes,
                ['id','name','en','hp','keyAttr','tradition','casting','skills','desc']))
        if subclasses is not None:
            js_parts.append(array_to_js('SUBCLASS_DB', subclasses,
                ['id','class_id','subclass_type','name_ko','name_en','summary']))
        if ancestries is not None:
            js_parts.append(array_to_js('ANCESTRIES', ancestries,
                ['id','name','en','hp','size','speed','boosts','flaws','traits','desc']))
        if backgrounds is not None:
            js_parts.append(array_to_js('BACKGROUNDS', backgrounds,
                ['id','name','en','boosts','skills','feat','desc']))
        if heritages is not None:
            js_parts.append(array_to_js('HERITAGE_DB', heritages,
                ['id','name_ko','name_en','ancestry','summary']))

        # TRAIT_DB, CONDITIONS, PROF_RANKS 등은 원본에서 보존
        # TRAIT_DB 추출
        trait_match = re.search(r'(const TRAIT_DB\s*=\s*\{.*?\};)', orig, re.DOTALL)
        if trait_match:
            js_parts.append(trait_match.group(1))

        if conditions_data is not None:
            js_parts.append(array_to_js('CONDITIONS_DATA', conditions_data,
                ['name','en','valued','max','desc']))

        if skills is not None:
            js_parts.append(array_to_js('SKILLS', skills,
                ['id','name','en','attr','isLore']))

        # SKILL_NAME_MAP 보존
        skill_map_match = re.search(r'(const SKILL_NAME_MAP\s*=\s*\{.*?\};)', orig, re.DOTALL)
        if skill_map_match:
            js_parts.append(skill_map_match.group(1))

        if actions is not None:
            js_parts.append(array_to_js('ACTION_DB', actions,
                ['id','cat','cat_label','name_ko','name_en','cost','traits','req_skill','req_rank','req_feat','summary']))

        # CONDITIONS, PROF_RANKS, state, modalType 등 보존
        tail_match = re.search(r'(const CONDITIONS\s*=\s*\[.*)', orig, re.DOTALL)
        if tail_match:
            js_parts.append(tail_match.group(1))

        final_js = '\n\n'.join(js_parts) + '\n'
        if not dry_run: cs_data_path.write_text(final_js, encoding='utf-8')
        print(f"    → cs_data.js {'(미리보기)' if dry_run else '✓'}")

    # ═══ class_features_db.js (부분 업데이트) ═══
    deity = read_sheet(wb, '신격_DEITY')
    class_features = read_dict_sheet(wb, '클래스요소_FEATURE')
    sub_features = read_dict_sheet(wb, '서브클래스요소_SUBFEATURE')
    auto_feats = read_dict_sheet(wb, '자동재주_AUTOFEAT')
    sub_auto_feats = read_dict_sheet(wb, '서브자동재주_SUBAUTOFEAT')
    auto_spells = read_dict_sheet(wb, '자동주문_AUTOSPELL')
    sub_auto_spells = read_dict_sheet(wb, '서브자동주문_SUBAUTOSPELL')

    cf_path = SCRIPT_DIR / 'class_features_db.js'
    cf_orig = cf_path.read_text(encoding='utf-8')

    # CLASS_PROF_TABLE, SUBCLASS_PROF_TABLE, PATRON_TRADITION, DIVINE_FONT_SLOTS → 보존 (수치 테이블은 수동 관리)
    cf_parts = []

    # 헤더 + PROF_TABLE 부분 보존 (첫 번째 CLASS_FEATURE_NAMES 이전까지)
    prof_end = cf_orig.find('var CLASS_FEATURE_NAMES')
    if prof_end > 0:
        cf_parts.append(cf_orig[:prof_end].rstrip())
    else:
        cf_parts.append(cf_orig.split('var CLASS_FEATURE_NAMES')[0].rstrip())

    if class_features is not None:
        cf_parts.append(dict_to_js('CLASS_FEATURE_NAMES', class_features))
    if auto_feats is not None:
        cf_parts.append(dict_to_js('CLASS_AUTO_FEATS', auto_feats))
    if sub_auto_feats is not None:
        cf_parts.append("// Subclass auto-granted feats")
        cf_parts.append(dict_to_js('SUBCLASS_AUTO_FEATS', sub_auto_feats))
    if auto_spells is not None:
        cf_parts.append(dict_to_js('CLASS_AUTO_SPELLS', auto_spells))
    if sub_auto_spells is not None:
        cf_parts.append(dict_to_js('SUBCLASS_AUTO_SPELLS', sub_auto_spells))
    if sub_features is not None:
        cf_parts.append(dict_to_js('SUBCLASS_FEATURE_NAMES', sub_features))

    if deity is not None:
        cf_parts.append(array_to_js('DEITY_DB', deity,
            ['id','name_ko','name_en','weapon','skill','sanctification','domains'], 'var'))

    # PATRON_TRADITION, DIVINE_FONT_SLOTS 보존
    patron_match = re.search(r'(var PATRON_TRADITION\s*=\s*\{.*?\};)', cf_orig, re.DOTALL)
    if patron_match:
        cf_parts.append(patron_match.group(1))
    font_match = re.search(r'(var DIVINE_FONT_SLOTS\s*=\s*\{.*?\};)', cf_orig, re.DOTALL)
    if font_match:
        cf_parts.append(font_match.group(1))

    cf_js = '\n\n'.join(cf_parts) + '\n'
    if not dry_run: cf_path.write_text(cf_js, encoding='utf-8')
    print(f"    → class_features_db.js {'(미리보기)' if dry_run else '✓'}")

    # ═══ cs_ui.js 내 DB 교체 (BARDING_DB, COMPANION_DB, FAMILIAR_ABILITIES) ═══
    companions = read_sheet(wb, '동물동료_COMPANION')
    bardings = read_sheet(wb, '마갑_BARDING')
    familiars = read_sheet(wb, '사역마능력_FAMILIAR')

    if any(x is not None for x in [companions, bardings, familiars]):
        ui_path = SCRIPT_DIR / 'cs_ui.js'
        ui_src = ui_path.read_text(encoding='utf-8')

        def replace_const_array(src, var_name, new_js):
            """소스 내 const/var VAR_NAME = [...]; 블록을 new_js로 교체"""
            pat = re.compile(r'(?:const|var|let)\s+' + re.escape(var_name) + r'\s*=\s*\[', re.MULTILINE)
            m = pat.search(src)
            if not m: return src
            start = m.start()
            depth = 0
            i = src.index('[', start)
            while i < len(src):
                if src[i] == '[': depth += 1
                elif src[i] == ']':
                    depth -= 1
                    if depth == 0: break
                i += 1
            end = i + 1
            if end < len(src) and src[end] == ';': end += 1
            return src[:start] + new_js + src[end:]

        if bardings is not None:
            new_js = array_to_js('BARDING_DB', bardings, ['name','ac','dex','check','speed','bulk','category'])
            ui_src = replace_const_array(ui_src, 'BARDING_DB', new_js)

        if companions is not None:
            new_js = array_to_js('COMPANION_DB', companions,
                ['id','name_ko','name_en','size','hp','str','dex','con','int','wis','cha',
                 'skill','senses','speed','speeds','mount','uncommon','attacks'])
            ui_src = replace_const_array(ui_src, 'COMPANION_DB', new_js)

        if familiars is not None:
            new_js = array_to_js('FAMILIAR_ABILITIES', familiars, ['id','name','en','desc'])
            ui_src = replace_const_array(ui_src, 'FAMILIAR_ABILITIES', new_js)

        if not dry_run: ui_path.write_text(ui_src, encoding='utf-8')
        print(f"    → cs_ui.js (BARDING_DB/COMPANION_DB/FAMILIAR_ABILITIES) {'(미리보기)' if dry_run else '✓'}")

    if dry_run:
        print(f"\n🔍 미리보기 — 파일 변경 없음")
    else:
        print(f"\n✅ DB 업데이트 완료! git push로 배포하세요.")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)
    cmd = sys.argv[1]
    if cmd == 'export': do_export()
    elif cmd == 'import': do_import('--dry' in sys.argv)
    else: print(f"사용법: python3 db_tools.py export | import [--dry]")
